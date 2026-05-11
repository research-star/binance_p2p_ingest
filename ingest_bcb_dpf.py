#!/usr/bin/env python3
"""
ingest_bcb_dpf.py — Scraper de tasas pasivas (Caja de Ahorro + DPF) del BCB.

Descarga el Excel más reciente de https://www.bcb.gob.bo/?q=tasas_interes,
parsea la hoja de tasas pasivas e inserta en bcb_dpf_rates (SQLite).

Uso:
    python3 ingest_bcb_dpf.py              # auto-detect latest
    python3 ingest_bcb_dpf.py --db otro.db  # DB custom
"""

import argparse
import re
import sqlite3
import sys
import tempfile
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

from config import NORMALIZED_DB

# ── Config ────────────────────────────────────────────────────────────────────

BCB_PAGE_URL = "https://www.bcb.gob.bo/?q=tasas_interes"
BCB_BASE_URL = "https://www.bcb.gob.bo/webdocs/tasas_interes/"
HEADERS = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:128.0) Gecko/20100101 Firefox/128.0"}
MAX_RETRIES = 3
RETRY_DELAY_S = 5

# ── Category detection ────────────────────────────────────────────────────────

# Known category headers in the BCB Excel (sheet2). These mark the start of a
# group of entities. The value in col A is ALL CAPS with no numeric data in the row.
CATEGORY_MAP = {
    "BANCOS MULTIPLES": "BANCOS MULTIPLES",
    "ENTIDADES ESPECIALIZADAS EN MICROFINANZAS": "MICROFINANZAS",
    "BANCOS PYME": "BANCOS PYME",
    "ENTIDADES FINANCIERAS DE VIVIENDA": "ENT. VIVIENDA",
    "COOPERATIVAS DE AHORRO Y CRÉDITO": "COOPERATIVAS",
    "COOPERATIVAS DE AHORRO Y CREDITO": "COOPERATIVAS",
    "COOPERATIVAS": "COOPERATIVAS",
    "INSTITUCIONES FINANCIERAS DE DESARROLLO": "IFD",
}

# Strings that appear in rows we should STOP parsing at (footer section)
STOP_MARKERS = {
    "TASAS DE INTERÉS DE LOS VALORES",
    "TASAS  DE  INTERÉS  DE  LOS  VALORES",
    "TASAS DE INTERES DE LOS VALORES",
    "DETALLE",
    "BCB REMESAS",
    "BCB DIRECTO",
    "PROMEDIOS PONDERADOS POR MONTO",
}

# Header/label strings to skip (not entities, not categories)
SKIP_PATTERNS = re.compile(
    r'^(entidades?|tasa|plazo|caja de ahorro|moneda|depositos?|'
    r'informaci[oó]n|tasas? (pasivas?|activas?|de inter[eé]s|interbancarias?)|'
    r'promedio|vigencia|fuente|mayor|mn|me|ufv|mvdol?|'
    r'empresarial|pyme|micro-cr[eé]dito|consumo|vivienda|'
    r'banco central de bolivia|semana|\*)',
    re.IGNORECASE
)

# Column mapping for sheet2 (0-indexed from col A)
# Col A=0: Entidad, B=1: CA BOB, C-J=2-9: DPF BOB, K=10: CA USD, L-S=11-18: DPF USD, T=19: UFV, U=20: MVDOL
COL_CA_BOB = 1
COL_DPF_BOB_START = 2   # cols 2..9
COL_CA_USD = 10
COL_DPF_USD_START = 11  # cols 11..18
COL_UFV = 19
COL_MVDOL = 20

DPF_PLAZOS = [30, 60, 90, 180, 360, 720, 1080, -1]  # -1 = "Mayor"

# ── Schema ────────────────────────────────────────────────────────────────────

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS bcb_dpf_rates (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  fetch_date TEXT NOT NULL,
  report_date TEXT NOT NULL,
  entidad TEXT NOT NULL,
  moneda TEXT NOT NULL,
  producto TEXT NOT NULL,
  plazo INTEGER,
  tasa REAL,
  categoria TEXT,
  UNIQUE(report_date, entidad, moneda, producto, plazo)
);
"""

CREATE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_dpf_report_date ON bcb_dpf_rates(report_date);",
    "CREATE INDEX IF NOT EXISTS idx_dpf_entidad ON bcb_dpf_rates(entidad);",
    "CREATE INDEX IF NOT EXISTS idx_dpf_categoria ON bcb_dpf_rates(categoria);",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def init_table(conn: sqlite3.Connection):
    conn.execute(CREATE_TABLE_SQL)
    # Add categoria column if upgrading from old schema
    try:
        conn.execute("ALTER TABLE bcb_dpf_rates ADD COLUMN categoria TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists
    for idx in CREATE_INDEXES_SQL:
        conn.execute(idx)
    conn.commit()


def fetch_with_retry(url: str, timeout: int = 60) -> requests.Response:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout)
            r.raise_for_status()
            return r
        except (requests.RequestException, requests.HTTPError) as e:
            if attempt == MAX_RETRIES:
                raise
            print(f"  Retry {attempt}/{MAX_RETRIES} for {url}: {e}", file=sys.stderr)
            time.sleep(RETRY_DELAY_S)


def find_latest_td_link(html: str) -> tuple[str, str]:
    """Find the most recent TD_DD%20MM%20YYYY.xlsx link. Returns (url, date YYYY-MM-DD)."""
    pattern = re.compile(
        r'href=["\']([^"\']*?/?(TD_(\d{2})%20(\d{2})%20(\d{4})(?:\s*\(\d+\))?\.xlsx))["\']',
        re.IGNORECASE
    )
    matches = []
    for m in pattern.finditer(html):
        href = m.group(1)
        day, month, year = int(m.group(3)), int(m.group(4)), int(m.group(5))
        try:
            date = datetime(year, month, day)
            date_str = date.strftime("%Y-%m-%d")
            matches.append((date, date_str, href))
        except ValueError:
            continue

    if not matches:
        raise RuntimeError("No TD_*.xlsx links found on BCB page")

    matches.sort(key=lambda x: x[0], reverse=True)
    _, date_str, href = matches[0]

    if href.startswith("http"):
        url = href
    elif href.startswith("/"):
        url = "https://www.bcb.gob.bo" + href
    else:
        url = BCB_BASE_URL + href.split("/")[-1]

    return url, date_str


def _is_category_row(cell_a_text: str) -> str | None:
    """Check if this row is a category header. Returns canonical category name or None."""
    upper = cell_a_text.upper().strip()
    for key, canonical in CATEGORY_MAP.items():
        if upper == key:
            return canonical
    return None


def _should_stop(cell_a_text: str) -> bool:
    """Check if we've hit the footer section and should stop parsing."""
    upper = cell_a_text.upper().strip()
    for marker in STOP_MARKERS:
        if marker in upper:
            return True
    return False


def _should_skip(cell_a_text: str) -> bool:
    """Check if this row is a header/label that should be skipped."""
    return bool(SKIP_PATTERNS.match(cell_a_text.strip()))


def _row_has_rate_data(row: tuple, min_col: int = 1, max_col: int = 21) -> bool:
    """Check if a row has at least one numeric value in the rate columns that looks like a real rate (0-100)."""
    for i in range(min_col, min(max_col, len(row))):
        v = row[i]
        if v is not None and isinstance(v, (int, float)) and v != 0 and v < 100:
            return True
    return False


def parse_excel(filepath: str) -> list[tuple]:
    """Parse sheet2 (TASAS PASIVAS) with category-aware iteration.

    Returns list of (entidad, moneda, producto, plazo, tasa, categoria).
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)

    # Sheet 2 is the passive rates sheet
    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        ws = wb.active
    else:
        ws = wb[sheet_names[1]]

    rows_data = []
    current_category = None
    parsing_started = False  # Wait until we hit first category

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # Skip header rows (rows 1-11 are title/headers/column labels)
        if row_idx <= 11:
            continue

        # Get cell A value
        cell_a = row[0] if len(row) > 0 else None
        if cell_a is None or not isinstance(cell_a, str) or not cell_a.strip():
            continue

        cell_a_text = cell_a.strip()

        # Check if we should stop (BCB footer section)
        if _should_stop(cell_a_text):
            break

        # Check if this is a category header
        cat = _is_category_row(cell_a_text)
        if cat is not None:
            current_category = cat
            parsing_started = True
            continue

        # Don't parse until we've seen the first category
        if not parsing_started:
            continue

        # Skip known header/label patterns
        if _should_skip(cell_a_text):
            continue

        # Skip rows that don't have any reasonable rate data (values 0-100)
        if not _row_has_rate_data(row):
            continue

        # This is an entity data row — extract rates
        entidad = cell_a_text

        def get_val(col_idx):
            if col_idx >= len(row):
                return None
            v = row[col_idx]
            if v is None or not isinstance(v, (int, float)):
                return None
            if v == 0:
                return None  # Zero = bank doesn't offer this product
            if v >= 100:
                return None  # Likely not a rate (could be a plazo or other number)
            return round(float(v), 4)

        # Caja de Ahorro BOB
        ca_bob = get_val(COL_CA_BOB)
        if ca_bob is not None:
            rows_data.append((entidad, "BOB", "CAJA_AHORRO", None, ca_bob, current_category))

        # DPF BOB (8 plazos)
        for i, plazo in enumerate(DPF_PLAZOS):
            val = get_val(COL_DPF_BOB_START + i)
            if val is not None:
                rows_data.append((entidad, "BOB", "DPF", plazo, val, current_category))

        # Caja de Ahorro USD
        ca_usd = get_val(COL_CA_USD)
        if ca_usd is not None:
            rows_data.append((entidad, "USD", "CAJA_AHORRO", None, ca_usd, current_category))

        # DPF USD (8 plazos)
        for i, plazo in enumerate(DPF_PLAZOS):
            val = get_val(COL_DPF_USD_START + i)
            if val is not None:
                rows_data.append((entidad, "USD", "DPF", plazo, val, current_category))

        # UFV
        ufv = get_val(COL_UFV)
        if ufv is not None:
            rows_data.append((entidad, "UFV", "DPF", None, ufv, current_category))

        # MVDOL
        mvdol = get_val(COL_MVDOL)
        if mvdol is not None:
            rows_data.append((entidad, "MVDOL", "DPF", None, mvdol, current_category))

    wb.close()
    return rows_data


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    t0 = time.time()

    parser = argparse.ArgumentParser(description="Scraper tasas pasivas BCB (DPF + Caja de Ahorro)")
    parser.add_argument("--db", type=Path, default=NORMALIZED_DB,
                        help=f"SQLite database (default: {NORMALIZED_DB})")
    args = parser.parse_args()

    # 1. Connect to DB and ensure table exists
    conn = sqlite3.connect(args.db)
    init_table(conn)

    # 2. Fetch BCB page
    try:
        resp = fetch_with_retry(BCB_PAGE_URL, timeout=30)
        html = resp.text
    except Exception as e:
        print(f"[bcb_dpf] mode=error stage=html_fetch detail={e}")
        sys.exit(1)

    # 3. Find latest Excel link
    try:
        xlsx_url, report_date = find_latest_td_link(html)
    except RuntimeError as e:
        print(f"[bcb_dpf] mode=error stage=html_parse detail={e}")
        sys.exit(1)

    # 4. Check if already processed
    existing = conn.execute(
        "SELECT 1 FROM bcb_dpf_rates WHERE report_date=? LIMIT 1", (report_date,)
    ).fetchone()
    if existing:
        print(f"[bcb_dpf] mode=skip reason=already_processed date={report_date}")
        conn.close()
        sys.exit(0)

    # 5. Download Excel
    try:
        print(f"  Downloading: {xlsx_url}")
        resp = fetch_with_retry(xlsx_url, timeout=60)
    except Exception as e:
        print(f"[bcb_dpf] mode=error stage=download detail={e}")
        conn.close()
        sys.exit(1)

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(resp.content)
    tmp.close()

    # 6. Parse
    try:
        rows_data = parse_excel(tmp.name)
    except Exception as e:
        print(f"[bcb_dpf] mode=error stage=parse detail={e}")
        conn.close()
        Path(tmp.name).unlink(missing_ok=True)
        sys.exit(1)
    finally:
        Path(tmp.name).unlink(missing_ok=True)

    if not rows_data:
        print(f"[bcb_dpf] mode=error stage=parse detail=no_data_rows_extracted")
        conn.close()
        sys.exit(1)

    # 7. Insert
    fetch_date = datetime.now(timezone.utc).isoformat()
    inserted = 0
    try:
        for entidad, moneda, producto, plazo, tasa, categoria in rows_data:
            conn.execute(
                "INSERT OR IGNORE INTO bcb_dpf_rates "
                "(fetch_date, report_date, entidad, moneda, producto, plazo, tasa, categoria) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (fetch_date, report_date, entidad, moneda, producto, plazo, tasa, categoria)
            )
            inserted += 1
        conn.commit()
    except Exception as e:
        print(f"[bcb_dpf] mode=error stage=insert detail={e}")
        conn.close()
        sys.exit(1)

    conn.close()
    duration = time.time() - t0
    print(f"[bcb_dpf] mode=ok report_date={report_date} rows_inserted={inserted} duration_s={duration:.2f}")


if __name__ == "__main__":
    main()
