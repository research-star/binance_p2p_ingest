#!/usr/bin/env python3
"""
ingest_embi.py — Scraper del EMBI Global Diversified Subindices via BCRD.

El Banco Central de la República Dominicana republica diariamente un Excel
con la serie histórica del spread EMBI por país (LATAM + agregados Global
y LATINO). Granularidad diaria, lag ~1 día hábil.

Fuente:
    https://bcrdgdcprod.blob.core.windows.net/documents/entorno-internacional/
        documents/Serie_Historica_Spread_del_EMBI.xlsx

Comportamiento:
  1. GET con If-None-Match basado en ETag persistido → 304 = no-op limpio.
  2. Si 200: snapshot a embi_audit/, persist ETag, parse Excel, UPSERT a
     embi_spreads (SQLite), rotación 7 días en audit folder.
  3. HC_EMBI: ping start / success / fail con body. Graceful si HC_EMBI no seteado.

Unidad: el Excel viene en percentage points (ej. Bolivia 2026-05-15 = 4.08).
Multiplicamos × 100 para guardar como bps (convención EMBI estándar = 408 bps).

N/A pre-debut: descartamos, no insertamos fila. La ausencia se infiere por
el LEFT JOIN del consumidor (dashboard.py).

Uso:
    python ingest_embi.py            # corrida normal (con ETag cache)
    python ingest_embi.py --force    # ignora ETag, re-descarga y re-procesa
    python ingest_embi.py --dry-run  # parsea pero no escribe SQLite ni audit
"""

import argparse
import os
import sqlite3
import sys
import tempfile
import time
import traceback
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

import openpyxl
import requests

from config import NORMALIZED_DB

# ── Constantes ────────────────────────────────────────────────────────────

EMBI_URL = (
    "https://bcrdgdcprod.blob.core.windows.net/documents/"
    "entorno-internacional/documents/Serie_Historica_Spread_del_EMBI.xlsx"
)
HEADERS = {"User-Agent": "Mozilla/5.0 (binance_p2p_ingest)"}
TIMEOUT_S = 60

AUDIT_DIR = Path("/opt/binance_p2p/embi_audit")
ETAG_FILE = AUDIT_DIR / ".last_etag"
ROTATE_DAYS = 7

# Mapeo header del Excel → nombre canónico (ASCII lowercase). Si BCRD cambia
# el header, el script falla con error claro en parse_workbook en vez de
# poblar columnas mal nombradas silenciosamente.
COUNTRY_MAP = {
    "Global":      "global",
    "LATINO":      "latino",
    "REP DOM":     "rep_dom",
    "Argentina":   "argentina",
    "Bolivia":     "bolivia",
    "Brasil":      "brasil",
    "Chile":       "chile",
    "Colombia":    "colombia",
    "Costa Rica":  "costa_rica",
    "Ecuador":     "ecuador",
    "El Salvador": "el_salvador",
    "Guatemala":   "guatemala",
    "Honduras":    "honduras",
    "México":      "mexico",
    "Paraguay":    "paraguay",
    "Perú":        "peru",
    "Panamá":      "panama",
    "Uruguay":     "uruguay",
    "Venezuela":   "venezuela",
    "RD-LATINO":   "rd_latino",
}

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS embi_spreads (
  fecha       TEXT NOT NULL,
  pais        TEXT NOT NULL,
  spread_bps  REAL NOT NULL,
  PRIMARY KEY (fecha, pais)
);
"""
CREATE_INDEX_SQL = (
    "CREATE INDEX IF NOT EXISTS idx_embi_pais_fecha "
    "ON embi_spreads (pais, fecha);"
)

EXPECTED_TITLE = "EMBI Global Diversified Subindices"
SHEET_NAME = "Serie Histórica"

HC_EMBI = os.environ.get("HC_EMBI", "").strip()


# ── Healthcheck ───────────────────────────────────────────────────────────

def hc_ping(suffix: str = "", body: str = ""):
    """Ping a healthchecks.io. No-op si HC_EMBI vacío. Falla silenciosa.

    suffix='' → success ping.
    suffix='start' → start ping (señala inicio de run).
    suffix='fail' → fail ping (alerta inmediata).
    body → POST con body (útil para resumen/stacktrace).
    """
    if not HC_EMBI:
        return
    url = f"https://hc-ping.com/{HC_EMBI}"
    if suffix:
        url = f"{url}/{suffix}"
    try:
        if body:
            requests.post(url, data=body.encode("utf-8"), timeout=10)
        else:
            requests.get(url, timeout=10)
    except Exception as e:
        print(f"[embi] WARN hc_ping_failed: {e}", file=sys.stderr)


# ── ETag persistido ───────────────────────────────────────────────────────

def read_etag() -> str | None:
    try:
        if ETAG_FILE.exists():
            return ETAG_FILE.read_text(encoding="utf-8").strip() or None
    except OSError:
        pass
    return None


def write_etag(etag: str):
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    ETAG_FILE.write_text(etag.strip(), encoding="utf-8")


# ── Audit folder rotation ─────────────────────────────────────────────────

def rotate_audit(days: int = ROTATE_DAYS) -> int:
    if not AUDIT_DIR.exists():
        return 0
    cutoff = time.time() - days * 86400
    removed = 0
    for f in AUDIT_DIR.glob("embi_*.xlsx"):
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
                removed += 1
        except OSError:
            pass
    return removed


# ── Excel parse ───────────────────────────────────────────────────────────

def parse_workbook(path: Path) -> list[tuple[str, str, float]]:
    """Unpivot del Excel → lista de tuplas (fecha_iso, pais_canonical, spread_bps).

    Falla con RuntimeError si el header no coincide con COUNTRY_MAP o si el
    título del sheet no es el esperado — protección contra cambios silenciosos
    del BCRD que poblarían el SQLite con basura.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"Sheet {SHEET_NAME!r} no encontrado. Sheets: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        iter_top = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        title_row = next(iter_top, ())
        header_row = next(iter_top, ())

        title = title_row[0] if title_row else None
        if title != EXPECTED_TITLE:
            raise RuntimeError(
                f"Title row mismatch. Expected {EXPECTED_TITLE!r}, got {title!r}"
            )
        if not header_row or header_row[0] != "Fecha":
            got = header_row[0] if header_row else None
            raise RuntimeError(f"Header[0] should be 'Fecha', got {got!r}")

        col_to_country = {}
        for idx, hdr in enumerate(header_row[1:], start=1):
            if hdr is None or hdr == "":
                continue
            if hdr not in COUNTRY_MAP:
                raise RuntimeError(
                    f"Header column {hdr!r} at col {idx} not in COUNTRY_MAP. "
                    f"BCRD may have added a country — update COUNTRY_MAP."
                )
            col_to_country[idx] = COUNTRY_MAP[hdr]

        if not col_to_country:
            raise RuntimeError("No country columns matched in header row")

        out: list[tuple[str, str, float]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            fecha = row[0] if row else None
            if not isinstance(fecha, datetime):
                continue
            fecha_iso = fecha.strftime("%Y-%m-%d")
            for col_idx, pais in col_to_country.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or val == "N/A" or val == "":
                    continue
                try:
                    pct = float(val)
                except (TypeError, ValueError):
                    continue
                bps = round(pct * 100, 4)
                out.append((fecha_iso, pais, bps))
        return out
    finally:
        wb.close()


# ── SQLite ────────────────────────────────────────────────────────────────

def init_table(conn: sqlite3.Connection):
    conn.execute(CREATE_TABLE_SQL)
    conn.execute(CREATE_INDEX_SQL)
    conn.commit()


def upsert(conn: sqlite3.Connection,
           rows: list[tuple[str, str, float]]) -> tuple[int, int, str]:
    """INSERT OR REPLACE — idempotente. Devuelve (n_filas, n_paises, last_fecha)."""
    conn.executemany(
        "INSERT OR REPLACE INTO embi_spreads (fecha, pais, spread_bps) "
        "VALUES (?, ?, ?)",
        rows
    )
    conn.commit()
    n_paises = len({r[1] for r in rows})
    last_fecha = max(r[0] for r in rows)
    return len(rows), n_paises, last_fecha


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest EMBI spreads from BCRD")
    parser.add_argument("--db", type=Path, default=NORMALIZED_DB,
                        help=f"SQLite database (default: {NORMALIZED_DB})")
    parser.add_argument("--force", action="store_true",
                        help="Ignorar ETag persistido y re-descargar")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parsea pero no escribe SQLite ni audit folder")
    args = parser.parse_args()

    t0 = time.time()
    hc_ping("start")

    try:
        AUDIT_DIR.mkdir(parents=True, exist_ok=True)
        audit_available = True
    except OSError as e:
        # En laptop dev (no es VPS), AUDIT_DIR fuera del repo puede fallar.
        # En ese caso desactivamos snapshot + ETag persistido y seguimos.
        print(f"[embi] WARN audit_dir_unavailable: {e}", file=sys.stderr)
        audit_available = False

    headers = dict(HEADERS)
    prev_etag = None if args.force else (read_etag() if audit_available else None)
    if prev_etag:
        headers["If-None-Match"] = prev_etag

    try:
        resp = requests.get(EMBI_URL, headers=headers, timeout=TIMEOUT_S)
    except Exception as e:
        msg = f"[embi] mode=error stage=fetch detail={e}"
        print(msg, file=sys.stderr)
        hc_ping("fail", body=msg)
        return 1

    if resp.status_code == 304:
        msg = (f"[embi] mode=ok stage=304 reason=etag_match "
               f"duration_s={time.time()-t0:.2f}")
        print(msg)
        hc_ping(body="304 not modified")
        return 0

    if resp.status_code != 200:
        msg = f"[embi] mode=error stage=fetch http={resp.status_code}"
        print(msg, file=sys.stderr)
        hc_ping("fail", body=msg)
        return 1

    new_etag = resp.headers.get("ETag", "").strip()
    content = resp.content

    # Snapshot a audit folder. Fecha BO ≈ UTC-4.
    today_bo = (datetime.now(timezone.utc) - timedelta(hours=4)).date().isoformat()
    audit_path: Path | None = None

    if audit_available and not args.dry_run:
        candidate = AUDIT_DIR / f"embi_{today_bo}.xlsx"
        try:
            candidate.write_bytes(content)
            audit_path = candidate
        except OSError as e:
            print(f"[embi] WARN audit_save_failed: {e}", file=sys.stderr)

    # Parse: usa el snapshot persistido si existe, sino tmp file.
    if audit_path is not None:
        parse_target = audit_path
        cleanup_target = False
    else:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(content)
        tmp.close()
        parse_target = Path(tmp.name)
        cleanup_target = True

    try:
        rows = parse_workbook(parse_target)
    except Exception as e:
        tb = traceback.format_exc()
        msg = f"[embi] mode=error stage=parse detail={e}"
        print(msg + "\n" + tb, file=sys.stderr)
        hc_ping("fail", body=msg + "\n" + tb)
        return 1
    finally:
        if cleanup_target:
            try:
                parse_target.unlink()
            except OSError:
                pass

    if not rows:
        msg = "[embi] mode=error stage=parse detail=zero_rows"
        print(msg, file=sys.stderr)
        hc_ping("fail", body=msg)
        return 1

    if args.dry_run:
        last_fecha = max(r[0] for r in rows)
        n_paises = len({r[1] for r in rows})
        msg = (f"[embi] mode=dry-run rows={len(rows)} countries={n_paises} "
               f"last_fecha={last_fecha} duration_s={time.time()-t0:.2f}")
        print(msg)
        hc_ping(body=msg)
        return 0

    # Persist ETag DESPUÉS del parse OK (no antes — evita "ETag avanzado, datos
    # no insertados" si el parse falla a mitad de camino).
    if new_etag and audit_available:
        try:
            write_etag(new_etag)
        except OSError as e:
            print(f"[embi] WARN etag_save_failed: {e}", file=sys.stderr)

    conn = sqlite3.connect(str(args.db))
    try:
        init_table(conn)
        n_rows, n_paises, last_fecha = upsert(conn, rows)
    finally:
        conn.close()

    removed = rotate_audit() if audit_available else 0

    msg = (f"[embi] mode=ok rows_upserted={n_rows} countries={n_paises} "
           f"last_fecha={last_fecha} etag={new_etag[:24]} "
           f"rotated={removed} duration_s={time.time()-t0:.2f}")
    print(msg)
    hc_ping(body=msg)
    return 0


if __name__ == "__main__":
    sys.exit(main())
