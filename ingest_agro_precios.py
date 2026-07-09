#!/usr/bin/env python3
"""
ingest_agro_precios.py — Precios agro para la tab Agro: FPMA + Pink Sheet + valor unitario.

Genera static/agro_precios.json combinando TRES fuentes públicas:

  1. FAO GIEWS FPMA (https://fpma.fao.org/giews/v4/global/price_module/api/v1/)
     Precios mayoristas/minoristas domésticos de Bolivia en BOB, mensuales.
     Se usa `price_value` (BOB nominal por unidad de la serie) — NUNCA
     `price_value_dollar` (usa el TC oficial ~6.9, inservible en Bolivia hoy).
     Cuando la unidad es de masa (quintal, libra, kg, arroba, cuartilla) se
     normaliza a Bs/kg multiplicando price_value x conversion_factor.
     Si la unidad NO es masa (ej. "0.9 Liter"), se emite con su unidad real.
     El inventario se re-fetchea live y la selección es por reglas (commodity +
     price_type + mercado), resiliente a uuids que cambien.

  2. Banco Mundial — Pink Sheet (CMO-Historical-Data-Monthly.xlsx)
     Precios internacionales de commodities en USD/t, mensuales, desde 2008-01.
     Café y azúcar vienen en $/kg → x1000 a USD/t.

  3. Valor unitario de exportación (FOB/t) — INE IneComex, vía
     static/agro_exportaciones.json (o el build de referencia de Trabajo previo
     si ese archivo aún no existe). Anual; 2026 es YTD (3 meses).

═══ REGLA DE PRECIOS DE DIEGO (CRÍTICA) ═══
Valor unitario FOB/ton SOLO para productos mono-partida homogéneos.
WHITELIST EXACTA: ['sesamo','chia','quinua','mani','castana','cafe'].
PROHIBIDO derivar precio de grupos mixtos (ej. soya y derivados =
aceite+torta+grano): sesgado y no representativo.
Toda serie lleva etiqueta de fuente visible.
═══════════════════════════════════════════

Fail-closed: si UNA fuente falla entera, se aborta SIN escribir output parcial
(el archivo previo queda intacto).

Uso:
    python ingest_agro_precios.py                # corrida normal
    python ingest_agro_precios.py --skip-fpma    # debug: sin FPMA
    python ingest_agro_precios.py --skip-pink    # debug: sin Pink Sheet
    python ingest_agro_precios.py --out otro.json --timeout 120
"""

import argparse
import io
import json
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ── Constantes ────────────────────────────────────────────────────────────

REPO_DIR = Path(__file__).parent
DEFAULT_OUT = REPO_DIR / "static" / "agro_precios.json"

FPMA_API = "https://fpma.fao.org/giews/v4/global/price_module/api/v1/"

PINK_PAGE = "https://www.worldbank.org/en/research/commodity-markets"
PINK_FALLBACK = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)
PINK_DESDE = "2008M01"  # recorte inferior de la serie internacional

# Input del valor unitario: primero el artefacto del repo (workstream B);
# fallback al build de referencia (contenido idéntico, solo lectura).
UV_REPO = REPO_DIR / "static" / "agro_exportaciones.json"
UV_FALLBACK = Path(r"C:\Dev\Trabajo previo\Comex\COMEX-Bolivia\build\granos_data.json")

# REGLA DE DIEGO: whitelist exacta de productos mono-partida homogéneos.
# NO agregar grupos mixtos (soya y derivados, azúcar+etanol, etc.).
UV_WHITELIST = ["sesamo", "chia", "quinua", "mani", "castana", "cafe"]

HEADERS = {"User-Agent": "Mozilla/5.0 (binance_p2p_ingest agro)"}

PRESUPUESTO_KB = 150  # tope de peso del JSON de salida

# Serie FPMA "muerta": último dato anterior a este mes (se descarta salvo
# que no haya alternativa viva del producto).
FPMA_MUERTA_ANTES = "2024-01"

# Selección FPMA — dominantes WHOLESALE: 'National Average' si existe, sino
# el mercado con serie viva más larga. (commodity_name exacto del API.)
FPMA_DOMINANTES = [
    # (commodity_name, slug producto, label corto base)
    ("Quinoa",                  "quinua",  "Quinua"),
    ("Maize (yellow)",          "maiz",    "Maíz amarillo"),
    ("Rice (first quality)",    "arroz",   "Arroz 1ra calidad"),
    ("Wheat (flour)",           "trigo",   "Harina de trigo"),
    ("Wheat (flour, imported)", "trigo",   "Harina de trigo importada"),
    ("Potatoes (Desirée)",      "papa",    "Papa Desirée"),
]

# Selección FPMA — RETAIL en mercados puntuales.
FPMA_RETAIL = [
    # (commodity_name, market_name, slug producto, label corto base)
    ("Sugar",         "Santa Cruz", "azucar",         "Azúcar"),
    ("Sugar",         "La Paz",     "azucar",         "Azúcar"),
    ("Bananas",       "La Paz",     "banana",         "Banana"),
    ("Sunflower oil", "La Paz",     "aceite_girasol", "Aceite de girasol"),
]

# Sanity de dirección del conversion_factor (factor DEBE llevar a Bs/kg,
# verificado con quinua National Average: 953.75 Bs/quintal x 0.022 ≈ 20.98
# Bs/kg en 2026-03). Rango esperado por unidad de masa; unidad fuera de este
# mapa con factor → WARN pero se acepta.
CF_RANGOS = {
    "Spanish quintal (46 kg)":   (0.019, 0.025),   # ~1/46
    "Kg":                        (0.999, 1.001),
    "Libra":                     (2.0,   2.3),     # ~1/0.46
    "Cuartilla (2.88 kg)":       (0.33,  0.36),    # ~1/2.88
    "Bolivian arroba (11.5 kg)": (0.08,  0.095),   # ~1/11.5
    "3 Libras":                  (0.70,  0.76),    # ~1/1.38
}

# Unidades NO-masa: no se fuerza kg, se emite la unidad real.
UNIDAD_NO_MASA = {
    "Liter":     "Bs/L",
    "0.9 Liter": "Bs/0.9 L",
    "1.8 Liter": "Bs/1.8 L",
    "1 unit":    "Bs/unidad",
}

# Pink Sheet: label exacto de la hoja (post-strip) → (id serie, slug, label display).
PINK_SERIES = [
    ("Soybeans",        "pink_soya",         "soya",       "Soya (grano)"),
    ("Soybean oil",     "pink_aceite_soya",  "aceite_soya", "Aceite de soya"),
    ("Soybean meal",    "pink_torta_soya",   "torta_soya", "Torta de soya"),
    ("Maize",           "pink_maiz",         "maiz",       "Maíz"),
    ("Rice, Thai 5%",   "pink_arroz",        "arroz",      "Arroz (Thai 5%)"),
    ("Wheat, US HRW",   "pink_trigo",        "trigo",      "Trigo (US HRW)"),
    ("Sugar, world",    "pink_azucar",       "azucar",     "Azúcar (mundo)"),
    ("Coffee, Arabica", "pink_cafe_arabica", "cafe",       "Café arábica"),
    ("Coffee, Robusta", "pink_cafe_robusta", "cafe",       "Café robusta"),
]

UV_LABELS = {
    "sesamo":  "Sésamo",
    "chia":    "Chía",
    "quinua":  "Quinua",
    "mani":    "Maní",
    "castana": "Castaña",
    "cafe":    "Café",
}

# Labels display de productos (solo los que llevan alguna serie).
PRODUCTO_LABELS = {
    "quinua": "Quinua", "maiz": "Maíz", "arroz": "Arroz",
    "trigo": "Trigo / harina", "azucar": "Azúcar", "banana": "Banana",
    "papa": "Papa", "aceite_girasol": "Aceite de girasol", "soya": "Soya",
    "aceite_soya": "Aceite de soya", "torta_soya": "Torta de soya",
    "cafe": "Café", "sesamo": "Sésamo", "chia": "Chía", "mani": "Maní",
    "castana": "Castaña",
}


# ── Helpers ───────────────────────────────────────────────────────────────

def ahora_iso() -> str:
    """Timestamp ISO UTC sin microsegundos."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def build_session() -> requests.Session:
    """Session con retries (429/5xx) para las descargas HTTP."""
    s = requests.Session()
    retry = Retry(total=3, backoff_factor=1.5,
                  status_forcelist=[429, 500, 502, 503, 504],
                  allowed_methods=["GET"])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update(HEADERS)
    return s


def slug_ascii(texto: str) -> str:
    """Slug ASCII lowercase para ids de serie (ej. 'La Paz' → 'lapaz')."""
    tabla = str.maketrans("áéíóúñü", "aeiounu")
    return re.sub(r"[^a-z0-9]", "", texto.lower().translate(tabla))


# ── Fuente 1: FAO GIEWS FPMA ─────────────────────────────────────────────

def fpma_inventario(session: requests.Session, timeout: int) -> list[dict]:
    """Descarga el inventario completo de series BOL (paginado)."""
    series = []
    url = f"{FPMA_API}FpmaSerie/?iso3_country_code=BOL&page_size=200"
    while url:
        r = session.get(url, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        series.extend(data["results"])
        url = data.get("next")
    if not series:
        raise RuntimeError("FPMA: inventario BOL vacío")
    return series


def _fin_serie(s: dict) -> str:
    """end_date mensual de la serie ('' si no tiene periodicity)."""
    per = s.get("periodicity") or []
    return max((p.get("end_date") or "" for p in per), default="")


def _inicio_serie(s: dict) -> str:
    per = s.get("periodicity") or []
    return min((p.get("start_date") or "9999" for p in per), default="9999")


def _viva(s: dict) -> bool:
    return _fin_serie(s)[:7] >= FPMA_MUERTA_ANTES


def _span_meses(s: dict) -> int:
    ini, fin = _inicio_serie(s), _fin_serie(s)
    if not fin or ini.startswith("9999"):
        return -1
    y0, m0 = int(ini[:4]), int(ini[5:7])
    y1, m1 = int(fin[:4]), int(fin[5:7])
    return (y1 - y0) * 12 + (m1 - m0)


def fpma_seleccionar(inventario: list[dict]) -> list[tuple[dict, str, str, str]]:
    """Aplica las reglas de selección → [(serie, slug, label, nivel), ...].

    Dominantes: WHOLESALE 'National Average' si existe, sino el mercado con
    la serie viva más larga. RETAIL: mercado puntual; si esa serie murió,
    fallback a la serie RETAIL viva más larga del producto (WARN).
    Series muertas (end < 2024-01) se descartan salvo que no haya viva.
    """
    seleccion = []

    for commodity, slug, label_base in FPMA_DOMINANTES:
        cands = [s for s in inventario
                 if s["commodity_name"] == commodity
                 and s["price_type"] == "WHOLESALE"]
        if not cands:
            raise RuntimeError(f"FPMA: sin series WHOLESALE para {commodity!r}")
        vivas = [s for s in cands if _viva(s)]
        pool = vivas or cands
        if not vivas:
            print(f"[agro-precios] WARN fpma sin serie viva para {commodity!r}, "
                  f"uso muerta", file=sys.stderr)
        nac = [s for s in pool if s["market_name"] == "National Average"]
        elegida = nac[0] if nac else max(pool, key=_span_meses)
        mercado_disp = ("prom. nacional" if elegida["market_name"] == "National Average"
                        else elegida["market_name"])
        seleccion.append((elegida, slug,
                          f"{label_base} mayorista ({mercado_disp})", "mayorista"))

    for commodity, market, slug, label_base in FPMA_RETAIL:
        cands = [s for s in inventario
                 if s["commodity_name"] == commodity
                 and s["price_type"] == "RETAIL"]
        if not cands:
            raise RuntimeError(f"FPMA: sin series RETAIL para {commodity!r}")
        exactas = [s for s in cands if s["market_name"] == market and _viva(s)]
        if exactas:
            elegida = exactas[0]
        else:
            vivas = [s for s in cands if _viva(s)]
            if not vivas:
                raise RuntimeError(
                    f"FPMA: sin serie RETAIL viva para {commodity!r} "
                    f"(pedida: {market})")
            elegida = max(vivas, key=_span_meses)
            print(f"[agro-precios] WARN fpma retail {commodity!r} en {market!r} "
                  f"no disponible/viva; fallback a {elegida['market_name']!r}",
                  file=sys.stderr)
        seleccion.append((elegida, slug,
                          f"{label_base} minorista ({elegida['market_name']})",
                          "minorista"))

    return seleccion


def fpma_datapoints(session: requests.Session, uuid: str, timeout: int) -> list[dict]:
    """Serie de precios mensual. OJO: barra final ANTES del query string es
    obligatoria — sin ella el API devuelve vacío."""
    url = f"{FPMA_API}FpmaSeriePrice/{uuid}/?periodicity=monthly"
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    dps = r.json().get("datapoints") or []
    if not dps:
        raise RuntimeError(f"FPMA: serie {uuid} sin datapoints")
    return dps


def fpma_construir(session: requests.Session, timeout: int) -> dict:
    """Construye las series FPMA → {id_serie: dict schema}."""
    inventario = fpma_inventario(session, timeout)
    print(f"[agro-precios] fpma inventario={len(inventario)} series BOL")
    seleccion = fpma_seleccionar(inventario)

    out = {}
    for serie, slug, label, nivel in seleccion:
        uuid = serie["uuid"]
        unidad_label = serie["measure_unit_label"]
        cf_serie = serie["conversion_factor"]

        # Normalización de unidad: masa → Bs/kg vía conversion_factor;
        # no-masa → unidad real, sin factor (regla: no forzar kg).
        if unidad_label in UNIDAD_NO_MASA:
            unidad, usa_factor = UNIDAD_NO_MASA[unidad_label], False
        elif cf_serie:
            unidad, usa_factor = "Bs/kg", True
            rango = CF_RANGOS.get(unidad_label)
            if rango and not (rango[0] <= cf_serie <= rango[1]):
                raise RuntimeError(
                    f"FPMA: conversion_factor {cf_serie} fuera de rango "
                    f"{rango} para unidad {unidad_label!r} ({uuid}) — "
                    f"verificar dirección del factor")
            if not rango:
                print(f"[agro-precios] WARN fpma unidad {unidad_label!r} sin "
                      f"rango de sanity (cf={cf_serie})", file=sys.stderr)
        else:
            raise RuntimeError(
                f"FPMA: unidad {unidad_label!r} sin conversion_factor ni "
                f"mapeo no-masa ({uuid})")

        dps = fpma_datapoints(session, uuid, timeout)
        puntos = []
        for dp in dps:
            pv = dp.get("price_value")
            if pv is None:
                continue  # hueco: se omite, no se rellena
            fecha = (dp.get("date") or "")[:7]
            if len(fecha) != 7:
                continue
            if usa_factor:
                cf = dp.get("conversion_factor") or cf_serie
                val = round(pv * cf, 2)
            else:
                val = round(pv, 2)
            puntos.append((fecha, val))
        if not puntos:
            raise RuntimeError(f"FPMA: serie {uuid} quedó sin puntos válidos")
        puntos.sort(key=lambda t: t[0])

        mercado = serie["market_name"]
        sid = "fpma_{}_{}_{}".format(
            slug if slug != "trigo" else
            ("trigo_harina_imp" if "imported" in serie["commodity_name"]
             else "trigo_harina"),
            "nacional" if mercado == "National Average" else slug_ascii(mercado),
            nivel)
        if sid in out:
            raise RuntimeError(f"FPMA: id de serie duplicado {sid}")
        out[sid] = {
            "producto": slug,
            "label": label,
            "fuente": "fpma",
            "fuente_label": f"FAO GIEWS FPMA ({serie['source_name']})",
            "mercado": mercado,
            "nivel": nivel,
            "unidad": unidad,
            "freq": "M",
            "fechas": [p[0] for p in puntos],
            "valores": [p[1] for p in puntos],
        }
        print(f"[agro-precios] fpma {sid}: {len(puntos)} puntos "
              f"{puntos[0][0]}..{puntos[-1][0]} [{unidad}]")
    return out


# ── Fuente 2: Banco Mundial Pink Sheet ───────────────────────────────────

def pink_descubrir_url(session: requests.Session, timeout: int) -> str:
    """Descubre la URL del XLSX mensual en la página de commodity markets;
    fallback a la URL conocida si el HTML cambió."""
    try:
        r = session.get(PINK_PAGE, timeout=timeout)
        r.raise_for_status()
        m = re.search(
            r'href="(https://[^"]+/related/CMO-Historical-Data-Monthly\.xlsx)"',
            r.text)
        if m:
            return m.group(1)
        print("[agro-precios] WARN pink: URL no hallada en la página, "
              "uso fallback", file=sys.stderr)
    except requests.RequestException as e:
        print(f"[agro-precios] WARN pink discovery falló ({e}), uso fallback",
              file=sys.stderr)
    return PINK_FALLBACK


def pink_construir(session: requests.Session, timeout: int) -> tuple[dict, str, str]:
    """Descarga y parsea la Pink Sheet → ({id: serie}, url_usada, release)."""
    url = pink_descubrir_url(session, timeout)
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True,
                                data_only=True)
    try:
        if "Monthly Prices" not in wb.sheetnames:
            raise RuntimeError(
                f"Pink Sheet: hoja 'Monthly Prices' ausente. "
                f"Hojas: {wb.sheetnames}")
        ws = wb["Monthly Prices"]
        filas = ws.iter_rows(values_only=True)

        # Filas 0-3: título/notas; la 3 (0-based) trae 'Updated on ...'.
        release = ""
        for i in range(4):
            row = next(filas, ())
            celda = str(row[0]) if row and row[0] is not None else ""
            if celda.startswith("Updated"):
                release = celda.strip()
        labels = [str(c).strip() if c is not None else "" for c in next(filas, ())]
        unidades = [str(c).strip() if c is not None else "" for c in next(filas, ())]

        col_por_label = {}
        for idx, lab in enumerate(labels):
            if lab:
                col_por_label.setdefault(lab, idx)

        series_cols = []
        for label_xlsx, sid, slug, display in PINK_SERIES:
            if label_xlsx not in col_por_label:
                raise RuntimeError(
                    f"Pink Sheet: label {label_xlsx!r} no encontrado en fila 4 "
                    f"— revisar si el BM renombró la columna")
            col = col_por_label[label_xlsx]
            unidad = unidades[col] if col < len(unidades) else ""
            if "/kg" in unidad:
                factor = 1000.0  # $/kg → USD/t
            elif "/mt" in unidad:
                factor = 1.0
            else:
                raise RuntimeError(
                    f"Pink Sheet: unidad inesperada {unidad!r} para "
                    f"{label_xlsx!r} (esperaba $/mt o $/kg)")
            series_cols.append((label_xlsx, sid, slug, display, col, factor))

        datos = {sid: [] for _, sid, *_ in series_cols}
        re_fecha = re.compile(r"^(\d{4})M(\d{2})$")
        for row in filas:
            fecha_raw = str(row[0]).strip() if row and row[0] is not None else ""
            m = re_fecha.match(fecha_raw)
            if not m or fecha_raw < PINK_DESDE:
                continue
            fecha = f"{m.group(1)}-{m.group(2)}"
            for _, sid, _, _, col, factor in series_cols:
                val = row[col] if col < len(row) else None
                try:
                    num = float(val)  # '…' y no-numéricos caen al except
                except (TypeError, ValueError):
                    continue
                datos[sid].append((fecha, round(num * factor, 2)))
    finally:
        wb.close()

    out = {}
    for label_xlsx, sid, slug, display, _, _ in series_cols:
        puntos = sorted(datos[sid], key=lambda t: t[0])
        if not puntos:
            raise RuntimeError(f"Pink Sheet: serie {label_xlsx!r} sin datos "
                               f"desde {PINK_DESDE}")
        out[sid] = {
            "producto": slug,
            "label": display,
            "fuente": "pink",
            "fuente_label": "Banco Mundial - Pink Sheet",
            "mercado": None,
            "nivel": "internacional",
            "unidad": "USD/t",
            "freq": "M",
            "fechas": [p[0] for p in puntos],
            "valores": [p[1] for p in puntos],
        }
        print(f"[agro-precios] pink {sid}: {len(puntos)} puntos "
              f"{puntos[0][0]}..{puntos[-1][0]}")
    return out, url, release


# ── Fuente 3: Valor unitario de exportación (whitelist) ──────────────────

def uv_construir() -> dict:
    """Valor unitario FOB/t anual por producto de la WHITELIST.

    REGLA DE DIEGO: solo productos mono-partida homogéneos (whitelist exacta);
    PROHIBIDO derivar precio de grupos mixtos (soya y derivados, etc.).
    """
    if UV_REPO.exists():
        path = UV_REPO
    elif UV_FALLBACK.exists():
        path = UV_FALLBACK
    else:
        raise RuntimeError(
            f"UV: no existe ni {UV_REPO} ni el fallback {UV_FALLBACK}")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    print(f"[agro-precios] uv input: {path}")

    years = data["meta"]["years"]
    ytd_years = set(data["meta"].get("ytdYears") or [])
    nacional = data["nacional"]

    out = {}
    for prod in UV_WHITELIST:
        if prod not in nacional:
            raise RuntimeError(f"UV: producto {prod!r} ausente en {path}")
        fob_a = nacional[prod]["fobA"]
        ton_a = nacional[prod]["tonA"]
        fechas, valores = [], []
        for i, anio in enumerate(years):
            fob = fob_a[i] if i < len(fob_a) else None
            ton = ton_a[i] if i < len(ton_a) else None
            if not ton or ton <= 0 or fob is None:
                continue  # sin tonelaje no hay valor unitario; hueco visible
            fechas.append(int(anio))
            valores.append(round(fob / ton, 2))
        if not fechas:
            raise RuntimeError(f"UV: producto {prod!r} sin años con tonA>0")
        out[f"uv_{prod}"] = {
            "producto": prod,
            "label": f"{UV_LABELS[prod]} valor unitario export.",
            "fuente": "uv",
            "fuente_label": "Valor unitario de exportación (FOB/t, INE)",
            "mercado": None,
            "nivel": "export",
            "unidad": "USD/t",
            "freq": "A",
            "fechas": fechas,
            "valores": valores,
            "ytd_ultimo": fechas[-1] in ytd_years,
        }
        print(f"[agro-precios] uv uv_{prod}: {len(fechas)} años "
              f"{fechas[0]}..{fechas[-1]} ytd={fechas[-1] in ytd_years}")
    return out


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera static/agro_precios.json (FPMA + Pink Sheet + UV)")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT,
                        help=f"Archivo de salida (default: {DEFAULT_OUT})")
    parser.add_argument("--skip-fpma", action="store_true",
                        help="Debug: omitir FAO GIEWS FPMA")
    parser.add_argument("--skip-pink", action="store_true",
                        help="Debug: omitir WB Pink Sheet")
    parser.add_argument("--timeout", type=int, default=60,
                        help="Timeout HTTP en segundos (default: 60)")
    args = parser.parse_args()

    t0 = time.time()
    session = build_session()
    series = {}
    meta_fuentes = {}

    # Fail-closed: cualquier excepción en una fuente aborta ANTES del write.
    try:
        if not args.skip_fpma:
            fpma_desc = ahora_iso()
            series.update(fpma_construir(session, args.timeout))
            meta_fuentes["fpma"] = {
                "api": FPMA_API,
                "descargado": fpma_desc,
                "nota": ("price_value nominal BOB; normalizado a Bs/kg via "
                         "conversion_factor cuando la unidad es masa"),
            }
    except Exception as e:
        print(f"[agro-precios] mode=error stage=fpma detail={e}", file=sys.stderr)
        return 1

    try:
        if not args.skip_pink:
            pink_desc = ahora_iso()
            pink_series, pink_url, pink_release = pink_construir(
                session, args.timeout)
            series.update(pink_series)
            meta_fuentes["pinksheet"] = {
                "url": pink_url,
                "release": pink_release,
                "descargado": pink_desc,
            }
    except Exception as e:
        print(f"[agro-precios] mode=error stage=pink detail={e}", file=sys.stderr)
        return 1

    try:
        series.update(uv_construir())
        meta_fuentes["unitvalue"] = {
            "fuente": "INE IneComex via agro_exportaciones.json",
            "regla": "solo productos mono-partida homogeneos (whitelist)",
            "whitelist": UV_WHITELIST,
        }
    except Exception as e:
        print(f"[agro-precios] mode=error stage=unitvalue detail={e}",
              file=sys.stderr)
        return 1

    # Índice de productos → series (orden de inserción: fpma, pink, uv).
    productos = {}
    for sid, s in series.items():
        slug = s["producto"]
        if slug not in productos:
            productos[slug] = {"label": PRODUCTO_LABELS[slug], "series": []}
        productos[slug]["series"].append(sid)

    doc = {
        "meta": {"generado": ahora_iso(), "fuentes": meta_fuentes},
        "productos": productos,
        "series": series,
    }
    payload = json.dumps(doc, ensure_ascii=False, separators=(",", ":"))
    peso_kb = len(payload.encode("utf-8")) / 1024
    if peso_kb > PRESUPUESTO_KB:
        print(f"[agro-precios] mode=error stage=budget peso={peso_kb:.1f}KB "
              f"> {PRESUPUESTO_KB}KB", file=sys.stderr)
        return 1

    args.out.parent.mkdir(parents=True, exist_ok=True)
    tmp = args.out.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(payload)
    tmp.replace(args.out)

    n_fpma = sum(1 for s in series.values() if s["fuente"] == "fpma")
    n_pink = sum(1 for s in series.values() if s["fuente"] == "pink")
    n_uv = sum(1 for s in series.values() if s["fuente"] == "uv")
    print(f"[agro-precios] mode=ok series={len(series)} "
          f"(fpma={n_fpma} pink={n_pink} uv={n_uv}) "
          f"productos={len(productos)} peso_kb={peso_kb:.1f} "
          f"out={args.out} duration_s={time.time()-t0:.1f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
