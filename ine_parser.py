"""
ine_parser.py — Adapters de parsing para los XLSX del INE Bolivia.

Cinco layouts soportados (dispatchados por `parse_cuadro()` según el `layout`
declarado en `config.INE_CUADROS`):

  - pib_trim_vertical         (PIB Trimestral 01.01.01, 01.01.04, 02.01.01)
  - pib_anual_wide            (PIB Anual Serie Histórica 06.01.01, 06.01.03)
  - ipc_nacional              (IPC Nacional general, 4 magnitudes en 4 hojas)
  - ipc_coicop_doubleheader   (IPC por División COICOP, 4 mag × 13 divisiones)
  - ipc_empalmada             (IPC Serie Histórica Empalmada 1937-presente)

Convenciones:
  - openpyxl crudo (sin pandas), espejando el patrón de `ingest_embi.py`.
  - Devolvemos dicts long-format (una fila por (period, dimension/indicador)).
  - El parser NO filtra preliminares ni huecos — filosofía del repo "no filtrar
    en origen". El consumidor decide.
  - Mojibake CP1252-encoded-as-UTF-8 lo reemplazamos a posteriori (best-effort).
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import openpyxl

# ─── Constantes compartidas ─────────────────────────────────────────────────

# Mojibake: el INE guarda algunas labels en latin-1 dentro de archivos
# nominalmente UTF-8 → openpyxl los devuelve con caracteres U+FFFD.
# Estos pares cubren los casos vistos en los samples 2026-05.
MOJI_FIX = [
    ("A�O", "AÑO"),
    ("ECON�MICA", "ECONÓMICA"),
    ("EXTRACCI�N", "EXTRACCIÓN"),
    ("PETR�LEO", "PETRÓLEO"),
    ("SEG�N", "SEGÚN"),
    ("HISTORICA", "HISTÓRICA"),
    ("BASICOS", "BÁSICOS"),
    ("b�sicos", "básicos"),
    ("ESTAD�STICA", "ESTADÍSTICA"),
    # Fallback por carácter: cuando ya no podemos resolver, reemplazo cosmético.
    ("�", "?"),
]

YEAR_PRELIM_RE = re.compile(
    # Tolerante: el INE a veces publica labels malformados como '2022p)' (sin
    # paréntesis abrir) — ver pib_trim_02_01_01 R171 release 2026-05.
    # Acepta '2024', '2024(p)', '2024 (p)', '2024p)', '2024(p', '2024 P'.
    r"^\s*(\d{4})\s*(\(?\s*p\s*\)?)?\s*$",
    re.IGNORECASE,
)

QUARTER_RE = re.compile(r"^\s*(I|II|III|IV)\s+Trimestre\s*$", re.IGNORECASE)
QUARTER_MAP = {"I": "Q1", "II": "Q2", "III": "Q3", "IV": "Q4"}

MESES_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


# ─── Utilidades de texto ────────────────────────────────────────────────────

def _fix_mojibake(s: str) -> str:
    for bad, good in MOJI_FIX:
        s = s.replace(bad, good)
    return s


def _norm_label(s: str | None) -> str:
    """Strip + collapse whitespace + fix mojibake. Conserva acentos UTF-8."""
    if s is None:
        return ""
    return _fix_mojibake(" ".join(str(s).split()))


def slugify(s: str) -> str:
    """ASCII lowercase + underscores. Idempotente."""
    s = _fix_mojibake(str(s))
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "unknown"


def _to_float(v) -> float | None:
    """Coerce a float si es numérico legítimo. Cualquier otra cosa (None, '',
    ' ', strings) → None. Aceptamos negativos legítimos (ej. SBI en PIB)."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


# ─── Adapter: PIB Trimestral (vertical) ─────────────────────────────────────

def parse_pib_trim(path: Path, cfg: dict) -> list[dict]:
    """Layout vertical: periodo en filas (5/año: anual + 4 trim), dimensiones
    en columnas. Header ~fila 10, unidad declarada en fila 9, footer en
    'Fuente:'. Sufijo '(p)' aplica a la celda de año.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        # Única hoja de datos. El cuadro id en la URL es 01.01.01 / 01.01.04 /
        # 02.01.01; la hoja se llama igual en el XLSX.
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        # Header en R10. Col 1 = 'PERIODO'/'PER�ODO', cols 2.. = dimensiones.
        # Toleramos shift por padding eventual; buscamos primer row cuya celda 1
        # empiece con 'PER' (PERIODO/PERÍODO).
        header_row = _find_header_row(ws, predicate=lambda v: isinstance(v, str)
                                      and _norm_label(v).upper().startswith("PER"))
        if header_row is None:
            raise RuntimeError(f"{path.name}: header row 'PERIODO' no encontrado")

        # Construir mapping col_idx -> dimension label/slug.
        dim_cols: list[tuple[int, str, str]] = []
        for col in range(2, ws.max_column + 1):
            label = _norm_label(ws.cell(header_row, col).value)
            if not label:
                continue
            dim_cols.append((col, label, slugify(label)))
        if not dim_cols:
            raise RuntimeError(f"{path.name}: 0 dimensiones en header row {header_row}")

        out: list[dict] = []
        current_year: int | None = None
        current_year_prelim = False

        for r in range(header_row + 1, ws.max_row + 1):
            label_raw = ws.cell(r, 1).value
            if label_raw is None:
                continue
            label = _norm_label(label_raw)
            if not label:
                continue
            # Fin del bloque de datos.
            if label.lower().startswith("fuente"):
                break

            # ¿Es una fila de año? '1990' o '2024(p)'.
            m_year = YEAR_PRELIM_RE.match(label)
            if m_year:
                current_year = int(m_year.group(1))
                current_year_prelim = bool(m_year.group(2))
                periodo = str(current_year)
            else:
                # ¿Es un trimestre? 'I Trimestre' .. 'IV Trimestre'.
                m_q = QUARTER_RE.match(label)
                if m_q and current_year is not None:
                    periodo = f"{current_year}-{QUARTER_MAP[m_q.group(1).upper()]}"
                else:
                    # Fila inesperada (footnote, header repetido, etc.) — skip.
                    continue

            for col, dim_label, dim_slug in dim_cols:
                v = _to_float(ws.cell(r, col).value)
                out.append({
                    "periodo": periodo,
                    "dimension": dim_slug,
                    "dimension_label": dim_label,
                    "valor": v,
                    "unidad": cfg["unit"],
                    "is_preliminary": 1 if current_year_prelim else 0,
                })
        return out
    finally:
        wb.close()


# ─── Adapter: PIB Anual Serie Histórica (wide) ──────────────────────────────

def parse_pib_anual_wide(path: Path, cfg: dict) -> list[dict]:
    """Layout WIDE: series (dimensiones) en filas, años en columnas.
    Cuadros 06.01.01 (actividad) y 06.01.03 (gasto). Cobertura 1980-presente.
    Header en R10, unidad en R9, data rows ~12-54 (variable según cuadro),
    footer 'Fuente:' al final. Quirks: filas separadoras vacías, label
    multi-línea (label sin valores → concatenar con fila siguiente),
    indentación por niveles, mojibake CP1252.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        # Header: fila con col 2 = 'ACTIVIDAD ECONOMICA' / 'COMPONENTES DEL GASTO'
        # y cols 3+ con años numéricos (con o sin '(p)').
        header_row = _find_header_row(ws, predicate=_anual_header_predicate)
        if header_row is None:
            raise RuntimeError(f"{path.name}: header row no encontrado en wide layout")

        # Mapping col_idx -> (year, is_prelim).
        year_cols: list[tuple[int, int, bool]] = []
        for col in range(3, ws.max_column + 1):
            v = ws.cell(header_row, col).value
            if v is None:
                continue
            m = YEAR_PRELIM_RE.match(str(v).strip())
            if m:
                year_cols.append((col, int(m.group(1)), bool(m.group(2))))
        if not year_cols:
            raise RuntimeError(f"{path.name}: 0 columnas de año en wide layout")

        out: list[dict] = []
        pending_label: str | None = None
        depth_of_pending: int = 0

        for r in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(r, 2).value
            if raw is None:
                # Separador vacío entre sectores → skip + reset pending.
                continue
            raw_str = str(raw)
            label = _norm_label(raw_str)
            if not label:
                continue
            if label.lower().startswith("fuente"):
                break
            if label.startswith("(p)"):  # nota footer "(p): Preliminar"
                continue

            # Indent depth: medirlo ANTES de strip.
            depth = len(raw_str) - len(raw_str.lstrip())

            # ¿Esta fila es label-only? (label presente, pero todas las celdas
            # de año son None o no-numéricas). Si sí, es continuación de un
            # label multilínea: guardar como prefix y mergeear con la próxima.
            row_vals = [_to_float(ws.cell(r, c).value) for c, _, _ in year_cols]
            if all(v is None for v in row_vals):
                pending_label = label
                depth_of_pending = depth
                continue

            if pending_label is not None:
                full_label = f"{pending_label} {label}"
                effective_depth = depth_of_pending
                pending_label = None
            else:
                full_label = label
                effective_depth = depth

            dim_slug = slugify(full_label)
            for (col, year, is_prelim), value in zip(year_cols, row_vals):
                out.append({
                    "periodo": str(year),
                    "dimension": dim_slug,
                    "dimension_label": full_label,
                    "depth": effective_depth,
                    "valor": value,
                    "unidad": cfg["unit"],
                    "is_preliminary": 1 if is_prelim else 0,
                })
        return out
    finally:
        wb.close()


def _anual_header_predicate(v) -> bool:
    """Identifica el header del wide layout. Distingue del título largo
    ('BOLIVIA: SERIE HISTORICA … SEGÚN ACTIVIDAD ECONÓMICA, 1980 - 2024')
    exigiendo que la celda sea CORTA — el header es un label compacto."""
    if not isinstance(v, str):
        return False
    nv = _norm_label(v).upper()
    if len(nv) > 50:
        return False
    # "COMPONENTE" cubre singular y plural; el 06.01.03 (gasto) usa 'COMPONENTE'.
    return ("ACTIVIDAD" in nv) or ("GASTO" in nv) or ("COMPONENTE" in nv)


# ─── Adapter: IPC Nacional + IPC Empalmada (single-band header) ─────────────

# Las 4 hojas de datos siguen el mismo patrón estructural. Difieren en:
#   - rango de años (2018-2026 nacional vs 1937-2026 empalmada)
#   - cuáles filas total trailing tienen (PROM. ANUAL, ACUMULADA — distinto
#     por hoja)
#   - sheet names mojibakeados según versión de openpyxl/Windows

_IPC_SHEET_PATTERNS = [
    (re.compile(r"1\.1.*[IÍ]NDICE", re.IGNORECASE | re.UNICODE), "indice",
     "pct_or_index", "indice_base_2016"),
    (re.compile(r"1\.2.*VAR.*MENSUAL", re.IGNORECASE),         "var_mensual",
     "pct", "pct_mensual"),
    (re.compile(r"1\.3.*ACUMULADA",   re.IGNORECASE),          "var_acumulada",
     "pct", "pct_acumulada"),
    (re.compile(r"1\.4.*12\s*MESES",  re.IGNORECASE),          "var_12m",
     "pct", "pct_12m"),
]


def _find_ipc_sheet(wb, pattern: re.Pattern) -> object | None:
    """Localizar la hoja por regex sobre sheetname.
    Fallback: si el nombre llega mojibakeado, normalizar antes de comparar."""
    for name in wb.sheetnames:
        if pattern.search(name) or pattern.search(_norm_label(name)):
            return wb[name]
    return None


def _ipc_single_header_parse(ws, indicador: str, unidad: str, base_year: str
                              ) -> list[dict]:
    """Layout común: R5 header 'MES | YEAR1 | YEAR2 | ...', R7..R18 = Enero..
    Diciembre, R19+ = totales. Filtrar trailing por whitelist de meses."""
    # Header de años — leemos toda la fila 5.
    year_cols: list[tuple[int, int]] = []
    for col in range(2, ws.max_column + 1):
        v = ws.cell(5, col).value
        if isinstance(v, (int, float)) and 1900 < int(v) < 2100:
            year_cols.append((col, int(v)))
        elif isinstance(v, str):
            m = YEAR_PRELIM_RE.match(v.strip())
            if m:
                year_cols.append((col, int(m.group(1))))
    if not year_cols:
        raise RuntimeError("IPC single-header: 0 columnas de año en R5")

    out: list[dict] = []
    # Iteramos sólo R7..R18 — los totales R19+ los descartamos por whitelist.
    for r in range(7, 19):
        mes_raw = ws.cell(r, 1).value
        mes = _norm_label(mes_raw).upper()
        month_num = MESES_ES.get(mes)
        if month_num is None:
            continue  # PROM. ANUAL, ACUMULADA, fila vacía, etc.
        for col, year in year_cols:
            v = _to_float(ws.cell(r, col).value)
            periodo = f"{year:04d}-{month_num:02d}"
            out.append({
                "periodo": periodo,
                "indicador": indicador,
                "indicador_label": indicador,
                "valor": v,
                "unidad": unidad,
                "base_year": base_year,
            })
    return out


def parse_ipc_nacional(path: Path, cfg: dict) -> list[dict]:
    """IPC Nacional general — 4 magnitudes en 4 hojas, indicadores planos."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        return _ipc_4sheets_parse(wb, cfg)
    finally:
        wb.close()


def parse_ipc_empalmada(path: Path, cfg: dict) -> list[dict]:
    """IPC Empalmada — mismo layout que IPC Nacional pero cobertura 1937-presente.
    El header R5 trae años como int en vez de str, y los totales R19/R20 difieren
    por hoja (la filtración por whitelist de meses los cubre uniformemente)."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        return _ipc_4sheets_parse(wb, cfg)
    finally:
        wb.close()


def _ipc_4sheets_parse(wb, cfg: dict) -> list[dict]:
    out: list[dict] = []
    for pattern, indicador, _kind, unidad in _IPC_SHEET_PATTERNS:
        ws = _find_ipc_sheet(wb, pattern)
        if ws is None:
            raise RuntimeError(f"IPC: no encontré hoja para patrón {pattern.pattern!r}")
        # base_year sólo aplica al índice nivel; en var % es None (no aplica
        # — una variación porcentual no tiene base year intrínseca).
        base_year = cfg.get("base_year") if indicador == "indice" else None
        out.extend(_ipc_single_header_parse(
            ws, indicador=indicador, unidad=unidad, base_year=base_year,
        ))
    return out


# ─── Adapter: IPC COICOP (double-band header año + mes) ─────────────────────

def parse_ipc_coicop(path: Path, cfg: dict) -> list[dict]:
    """IPC por División COICOP. Layout WIDE con doble header:
      R5: año (en cabecera mergeada cada 12 cols; openpyxl devuelve año sólo en
          la celda top-left del merge — propagamos a derecha)
      R6: mes (ENERO..DICIEMBRE)
      R8..R20: 13 filas de datos (división 0 = ÍNDICE GENERAL, 1-12 = divisiones COICOP)
      Cols: A=division_id (0..12), B=descripción, C..CY=valores (101 cells = 9
            años × 12 meses - 7 meses futuros)
      R21: 'Fuente:' footer
    4 hojas (1.1 índice / 1.2 var mensual / 1.3 acumulada / 1.4 12m).
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        out: list[dict] = []
        for pattern, metric, _kind, unidad in _IPC_SHEET_PATTERNS:
            ws = _find_ipc_sheet(wb, pattern)
            if ws is None:
                raise RuntimeError(
                    f"IPC COICOP: no encontré hoja {pattern.pattern!r}")
            # 1. Construir (col_idx -> (year, month)) usando R5+R6. Año se
            #    propaga hacia la derecha (merge cada 12 cols).
            col_to_ym: dict[int, tuple[int, int]] = {}
            current_year: int | None = None
            for col in range(3, ws.max_column + 1):
                y = ws.cell(5, col).value
                if isinstance(y, (int, float)) and 1900 < int(y) < 2100:
                    current_year = int(y)
                elif isinstance(y, str):
                    m = YEAR_PRELIM_RE.match(y.strip())
                    if m:
                        current_year = int(m.group(1))
                mes_raw = ws.cell(6, col).value
                if mes_raw is None or current_year is None:
                    continue
                month_num = MESES_ES.get(_norm_label(mes_raw).upper())
                if month_num is None:
                    continue
                col_to_ym[col] = (current_year, month_num)
            if not col_to_ym:
                raise RuntimeError(
                    f"IPC COICOP: 0 columnas (año,mes) en hoja {ws.title!r}")
            # 2. Filas de datos R8..R20. Validar division_id es int en col A.
            base_year = cfg.get("base_year") if metric == "indice" else None
            for r in range(8, 21):
                div_raw = ws.cell(r, 1).value
                if div_raw is None:
                    continue
                try:
                    div_id = int(div_raw)
                except (TypeError, ValueError):
                    continue
                div_label = _norm_label(ws.cell(r, 2).value)
                if not div_label:
                    continue
                div_slug = slugify(div_label)
                # Indicador compound. División 0 es "ÍNDICE GENERAL" (total
                # Bolivia, igual al cuadro ipc_nacional_general); usamos sufijo
                # '_total' para que la query LIKE '<metric>_d__%' la deje fuera
                # y el bar chart de 12 divisiones no la incluya como 13ra barra.
                if div_id == 0:
                    indicador = f"{metric}_total"
                else:
                    indicador = f"{metric}_{div_slug}"
                for col, (year, month_num) in col_to_ym.items():
                    v = _to_float(ws.cell(r, col).value)
                    periodo = f"{year:04d}-{month_num:02d}"
                    out.append({
                        "periodo": periodo,
                        "indicador": indicador,
                        "indicador_label": f"{metric} · {div_label}",
                        "valor": v,
                        "unidad": unidad,
                        "base_year": base_year,
                        "division_id": div_id,
                    })
        return out
    finally:
        wb.close()


# ─── Header-row finder (compartido) ─────────────────────────────────────────

def _find_header_row(ws, predicate, max_scan: int = 20) -> int | None:
    """Devuelve el índice de la primera fila ≤ max_scan donde alguna celda
    de las primeras 4 columnas satisface `predicate(value)`. None si no."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, 5):
            if predicate(ws.cell(r, c).value):
                return r
    return None


# ─── Dispatch ───────────────────────────────────────────────────────────────

LAYOUT_DISPATCH = {
    "pib_trim_vertical":         parse_pib_trim,
    "pib_anual_wide":            parse_pib_anual_wide,
    "ipc_nacional":              parse_ipc_nacional,
    "ipc_coicop_doubleheader":   parse_ipc_coicop,
    "ipc_empalmada":             parse_ipc_empalmada,
    # IPP reusa los adapters de IPC — la estructura del XLSX (4 hojas
    # 1.1-1.4, header single/double-band, base 2016=100) es idéntica;
    # sólo cambia la semántica del eje no-temporal (grandes grupos de
    # actividad vs divisiones COICOP). Los cuadros se discriminan por
    # `cuadro` namespaceado, no por layout.
    "ipp_nacional":              parse_ipc_nacional,
    "ipp_grandes_grupos":        parse_ipc_coicop,
}


def parse_cuadro(cuadro_id: str, path: Path, cfg: dict) -> list[dict]:
    layout = cfg["layout"]
    fn = LAYOUT_DISPATCH.get(layout)
    if fn is None:
        raise RuntimeError(f"{cuadro_id}: layout {layout!r} no soportado")
    return fn(path, cfg)
