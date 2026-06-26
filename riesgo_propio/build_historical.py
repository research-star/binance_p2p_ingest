#!/usr/bin/env python3
"""
build_historical.py - Historical Bolivia own-math sovereign spread ("Bolivia 2").

Reconstructs the FinanzasBo own-math riesgo-pais number BACK IN TIME using only
the data sources we actually have:

  * Bond terms .......... full historical Bolivia USD sovereign universe (incl.
                          the matured 4.875% 2022 and 5.95% 2023) -> BOLIVIA_BONDS.
  * UST par curves ...... US Treasury daily par curve, per-year CSV (free, exact).
  * Market spread level .. BCRD redistribution of the JPMorgan EMBI Global
                          Diversified Bolivia series (daily %, x100 = bp).

Per observation date d:
  1. basket B(d) = bonds outstanding (issue <= settle < maturity).
  2. zero curve Z(d) bootstrapped from that day's UST par curve.
  3. each bond's market yield ~ matched-tenor UST(d) + EMBI_spread(d); from that
     yield we get its dirty price; then we run OUR MV-weighted Z-spread over Z(d)
     on the actual outstanding cashflows.  -> own-math bp for date d.

The difference vs the raw EMBI number is the genuine methodology basis (par->zero
curve conversion across the ACTUAL outstanding basket + MV weighting). The live
edge is overridden with the validated own-math anchor computed from real prices.

NOTE: per-bond intraday historical PRICES are not freely available for these
144A/RegS sovereigns (DB/Stuttgart are live-only; Frankfurt API gated), so the
historical market LEVEL is anchored to EMBI. This is disclosed on the chart.
"""
from __future__ import annotations
import csv, io, json, os, ssl, sys, urllib.request
from datetime import date, timedelta

# --- reuse the validated riesgo_pais engine (do not modify it) ----------------
RP_SRC = r"C:\Users\RodrigoRosasGuzman\riesgo_pais\src"
sys.path.insert(0, RP_SRC)
import engine            # noqa: E402  Z-spread, ZeroCurve, interpolate, dirty_price_from_ytm
import bonds as bondmod  # noqa: E402  Bond, generate_cashflows

HERE = os.path.dirname(os.path.abspath(__file__))
CTX = ssl.create_default_context(); CTX.check_hostname = False; CTX.verify_mode = ssl.CERT_NONE

# Live anchor = the latest REAL price-driven own-math point recorded by
# live_bolivia.py (riesgo_propio_live.json). Falls back to the validated
# 2026-06-24 snapshot (434 bp) if the live store does not exist yet.
def _load_live_points():
    p = os.path.join(HERE, "riesgo_propio_live.json")
    if os.path.exists(p):
        try:
            h = json.load(open(p, encoding="utf-8"))
            if h:
                return sorted(h, key=lambda r: r["date"])
        except Exception:
            pass
    return []

_LIVE = _load_live_points()
if _LIVE:
    LIVE_ANCHOR_BP = float(_LIVE[-1]["bp"])
    LIVE_ANCHOR_DATE = date.fromisoformat(_LIVE[-1]["date"])
else:
    LIVE_ANCHOR_BP = 434.0
    LIVE_ANCHOR_DATE = date(2026, 6, 24)

# --- Bolivia USD sovereign universe (provisional; matured bonds confirmed by
#     the research pass before the full run). amount = USD millions outstanding.
def B(country, name, isin, issue, mat, coupon, amt, structure="bullet"):
    return bondmod.Bond(
        country=country, name=name, isin=isin,
        issue_date=date.fromisoformat(issue), maturity=date.fromisoformat(mat),
        coupon_schedule=[(date.fromisoformat(issue), coupon)], freq_months=6,
        amort_schedule=None, amount_outstanding_musd=amt, structure=structure,
    )

BOLIVIA_BONDS = [
    # name                       isin            issue        maturity     cpn    amtMM
    B("Bolivia","BOLIVIA 4.875% 2022","USP3779PAA28","2012-10-29","2022-10-22",4.875, 500),
    B("Bolivia","BOLIVIA 5.95% 2023", "USP3779PAB01","2013-08-08","2023-08-08",5.95,  500),
    B("Bolivia","BOLIVIA 4.50% 2028", "USP37878AC26","2017-03-20","2028-03-20",4.50, 1000),
    B("Bolivia","BOLIVIA 7.50% 2030", "USP37878AE81","2020-03-02","2030-03-02",7.50,  850),
    B("Bolivia","BOLIVIA 9.45% 2031", "USP37878AF56","2026-05-14","2031-05-14",9.45, 1000),
]

# ============================ data sources ====================================
def download_embi():
    """{date: spread_bp} for Bolivia from the BCRD workbook (cached locally)."""
    from openpyxl import load_workbook
    path = os.path.join(HERE, "embi_workbook.xlsx")
    if not os.path.exists(path):
        u = ("https://bcrdgdcprod.blob.core.windows.net/documents/entorno-internacional/"
             "documents/Serie_Historica_Spread_del_EMBI.xlsx")
        req = urllib.request.Request(u, headers={"User-Agent": "Mozilla/5.0"})
        open(path, "wb").write(urllib.request.urlopen(req, timeout=40, context=CTX).read())
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows[:15])
                 if r and any(isinstance(c, str) and "boliv" in c.lower() for c in r))
    bcol = next(j for j, c in enumerate(rows[hdr_i]) if isinstance(c, str) and "boliv" in c.lower())
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or r[0] is None or r[bcol] is None:
            continue
        v = r[bcol]
        if isinstance(v, str):          # "N/A" pre-debut
            continue
        d = r[0].date() if hasattr(r[0], "date") else r[0]
        out[d] = float(v) * 100.0       # percent -> bp
    return out

_UST_LABELS = {"1 Mo": 1/12, "2 Mo": 2/12, "3 Mo": 0.25, "4 Mo": 4/12, "6 Mo": 0.5,
               "1 Yr": 1, "2 Yr": 2, "3 Yr": 3, "5 Yr": 5, "7 Yr": 7,
               "10 Yr": 10, "20 Yr": 20, "30 Yr": 30}

def download_ust_year(yr):
    u = ("https://home.treasury.gov/resource-center/data-chart-center/interest-rates/"
         "daily-treasury-rates.csv/{0}/all?field_tdr_date_value={0}"
         "&type=daily_treasury_yield_curve&page&_format=csv").format(yr)
    req = urllib.request.Request(u, headers={"User-Agent": "Mozilla/5.0"})
    rows = list(csv.DictReader(io.StringIO(
        urllib.request.urlopen(req, timeout=30, context=CTX).read().decode())))
    out = {}
    for row in rows:
        try:
            m, dd, yy = row["Date"].split("/"); d = date(int(yy), int(m), int(dd))
        except Exception:
            continue
        curve = {}
        for lbl, ten in _UST_LABELS.items():
            v = row.get(lbl)
            if v not in (None, "", "N/A"):
                try: curve[ten] = float(v)
                except ValueError: pass
        if curve:
            out[d] = curve
    return out

class USTHistory:
    def __init__(self): self.byyear = {}; self.sorted = {}
    def _ensure(self, yr):
        if yr not in self.byyear:
            try: self.byyear[yr] = download_ust_year(yr)
            except Exception: self.byyear[yr] = {}
            self.sorted[yr] = sorted(self.byyear[yr])
    def curve_on(self, d):
        """Nearest prior business-day par curve (search up to ~10 days back, prev year ok)."""
        for back in range(0, 12):
            dd = d - timedelta(days=back); self._ensure(dd.year)
            if dd in self.byyear.get(dd.year, {}):
                return self.byyear[dd.year][dd]
        return None

# ============================ reconstruction ==================================
def matched_ust(curve, t_years):
    return engine.interpolate(curve, t_years)            # percent

def own_math_bp(d, embi_bp, curve):
    """MV-weighted own-math Z-spread (bp) for the basket outstanding on date d."""
    settle = d + timedelta(days=2)
    zc = engine.ZeroCurve(curve)
    s_dec = embi_bp / 10000.0
    num = den = 0.0
    n = 0
    for b in BOLIVIA_BONDS:
        if not (b.issue_date <= settle < b.maturity):
            continue
        cfs = bondmod.generate_cashflows(b)
        t = engine.yearfrac_30_360(settle, b.maturity)
        ytm_semi = matched_ust(curve, t) / 100.0 + s_dec        # market yield proxy
        dirty = engine.dirty_price_from_ytm(ytm_semi, settle, cfs)
        if dirty <= 0:
            continue
        zs = engine.zspread(dirty, settle, cfs, zc)             # OUR own-math spread
        mv = b.amount_outstanding_musd * dirty / 100.0          # market-value weight
        num += mv * zs; den += mv; n += 1
    if den == 0:
        return None, 0
    return num / den * 10000.0, n

def fridays(d0, d1):
    d = d0 + timedelta((4 - d0.weekday()) % 7)
    while d <= d1:
        yield d; d += timedelta(days=7)

def main():
    embi = download_embi()
    edates = sorted(embi)
    start, end = edates[0], edates[-1]
    print(f"EMBI Bolivia: {len(embi)} pts {start}..{end}")
    ust = USTHistory()
    fechas, serie, embi_line, nbonds = [], [], [], []
    embi_sorted = edates
    import bisect
    def embi_on(d):
        i = bisect.bisect_right(embi_sorted, d) - 1
        return embi[embi_sorted[i]] if i >= 0 else None
    for d in edates:                      # DAILY cadence (match EMBI) -> smooth line
        e = embi[d]
        curve = ust.curve_on(d)
        if not curve: continue
        bp, n = own_math_bp(d, e, curve)
        if bp is None: continue
        fechas.append(d.isoformat()); serie.append(bp)
        embi_line.append(round(e, 1)); nbonds.append(n)
    # Level-anchor the whole series to the validated real-price own-math number
    # (434 bp on 2026-06-24). This applies the structural own-math vs EMBI basis
    # measured from REAL current bond prices as a smooth shift -> no end cliff.
    if fechas:
        ai = min(range(len(fechas)),
                 key=lambda i: abs(date.fromisoformat(fechas[i]) - LIVE_ANCHOR_DATE))
        basis = LIVE_ANCHOR_BP - serie[ai]
        serie = [round(v + basis, 1) for v in serie]
    out = {"fechas": fechas, "bolivia_propio": serie, "bolivia_embi": embi_line,
           "n_bonds": nbonds, "live_anchor_bp": LIVE_ANCHOR_BP,
           "live_anchor_date": LIVE_ANCHOR_DATE.isoformat(),
           "live_points": [{"ts": r.get("ts", r["date"]), "bp": r["bp"]} for r in _LIVE]}
    json.dump(out, open(os.path.join(HERE, "riesgo_propio.json"), "w"), indent=0)
    print(f"wrote riesgo_propio.json: {len(fechas)} weekly pts {fechas[0]}..{fechas[-1]}")
    # quick sanity: last 4 points and basket-size transitions
    print("last pts (fecha, propio, embi, n):")
    for i in range(max(0, len(fechas)-4), len(fechas)):
        print("  ", fechas[i], serie[i], embi_line[i], nbonds[i])
    seen = None
    print("basket-size changes:")
    for f, n in zip(fechas, nbonds):
        if n != seen: print("  ", f, "-> n_bonds", n); seen = n

if __name__ == "__main__":
    main()
